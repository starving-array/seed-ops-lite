import contextlib
import csv
import io
import json
import uuid
from typing import Any

from fastapi import Depends, File, Form, HTTPException, UploadFile, status

from app.api.deps import get_runtime_provider
from app.api.endpoints.schema.helpers import DEFAULT_SCHEMA, RuntimeProviderType
from app.platform.container import get_persistence_provider
from app.platform.persistence.interfaces import PersistenceProvider
from app.platform.persistence.resolver import ProjectResolver
from app.schemas.schema_design import (
    ColumnModel,
    RelationshipModel,
    SchemaModel,
    TableModel,
)


async def load_schema(
    db: PersistenceProvider = Depends(get_persistence_provider),
) -> SchemaModel:
    """Loads the currently saved active schema state from SQLite."""
    project_id = ProjectResolver.get_active_project_id()
    try:
        schema_dict = await db.get_active_schema(project_id=project_id)
        if not schema_dict:
            # Verify and ensure default project exists
            if not await db.get_project(project_id):
                await db.create_project(project_id, "Default Project")

            schema_dict = await db.save_schema(
                project_id=project_id,
                version=1,
                tables=DEFAULT_SCHEMA["tables"],
                relationships=DEFAULT_SCHEMA.get("relationships", []),
            )

        return SchemaModel(
            tables=schema_dict["tables"],
            relationships=schema_dict.get("relationships", []),
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to load schema: {exc}",
        ) from exc


async def save_schema(
    schema: SchemaModel,
    db: PersistenceProvider = Depends(get_persistence_provider),
    runtime: RuntimeProviderType = Depends(get_runtime_provider),
) -> dict[str, str]:
    """Saves the current schema state to SQLite."""
    from app.seeder.column_guide import invalidate_column_guide

    project_id = ProjectResolver.get_active_project_id()
    try:
        # Determine next increment version num
        current_schema = await db.get_active_schema(project_id=project_id)
        next_version = 1
        if current_schema:
            next_version = current_schema["version"] + 1

        tables_dict = [t.model_dump(by_alias=True) for t in schema.tables]
        relationships_dict = [r.model_dump(by_alias=True) for r in schema.relationships]

        await db.save_schema(
            project_id=project_id,
            version=next_version,
            tables=tables_dict,
            relationships=relationships_dict,
        )

        # Invalidate column guide — schema changed
        with contextlib.suppress(Exception):
            await invalidate_column_guide(schema, runtime, db)

        return {"status": "success", "message": "Schema saved successfully"}
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save schema: {exc}",
        ) from exc


async def put_schema(
    schema: SchemaModel,
    db: PersistenceProvider = Depends(get_persistence_provider),
    runtime: RuntimeProviderType = Depends(get_runtime_provider),
) -> dict[str, str]:
    """Updates/Saves the current schema state to SQLite."""
    return await save_schema(schema, db, runtime)


async def delete_schema(
    db: PersistenceProvider = Depends(get_persistence_provider),
    runtime: RuntimeProviderType = Depends(get_runtime_provider),
) -> dict[str, str]:
    """Deactivates/deletes the active schema state in SQLite."""
    from app.seeder.column_guide import invalidate_column_guide

    project_id = ProjectResolver.get_active_project_id()
    try:
        # Fetch current schema before deactivating to get the fingerprint
        current = await db.get_active_schema(project_id=project_id)
        if current:
            from app.schemas.schema_design import SchemaModel

            schema = SchemaModel(
                tables=current.get("tables", []),
                relationships=current.get("relationships", []),
            )
            with contextlib.suppress(Exception):
                await invalidate_column_guide(schema, runtime, db)

        await db.deactivate_schema(project_id=project_id)
        return {"status": "success", "message": "Schema deleted successfully"}
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete schema: {exc}",
        ) from exc


async def import_schema(
    content: str | None = Form(default=None),
    file_type: str | None = Form(default=None),
    file: UploadFile | None = File(default=None),  # noqa: B008
    db: PersistenceProvider = Depends(get_persistence_provider),
    runtime: RuntimeProviderType = Depends(get_runtime_provider),
) -> SchemaModel:
    """Import schema from a SQL, DDL, CSV, JSON, TXT, or Excel file and save to the active project."""
    project_id = ProjectResolver.get_active_project_id()

    file_bytes = b""
    filename = ""
    if file:
        file_bytes = await file.read()
        filename = file.filename or ""
        if not file_type:
            # resolve type from filename extension
            ext = filename.split(".")[-1].lower() if "." in filename else ""
            file_type = ext

    if not file_type:
        file_type = "sql"

    file_type = file_type.lower()

    # Convert bytes to string if needed
    text_content = ""
    if file_bytes:
        try:
            text_content = file_bytes.decode("utf-8")
        except Exception:
            text_content = ""
    elif content:
        text_content = content

    tables = []
    relationships = []

    try:
        if file_type in ("sql", "ddl"):
            # Use DDLValidator
            from app.validation.ddl_validator import DDLValidator

            validator = DDLValidator()
            validator.validate(text_content)
            parsed_tables = validator.last_parsed_tables

            table_name_to_id = {}
            col_name_to_id = {}  # (table_name, col_name) -> id

            for tbl_name, tbl_def in parsed_tables.items():
                t_id = str(uuid.uuid4())[:8]
                table_name_to_id[tbl_name.lower()] = t_id

                cols = []
                for col_name, col_def in tbl_def.columns.items():
                    c_id = "c_" + str(uuid.uuid4())[:6]
                    col_name_to_id[(tbl_name.lower(), col_name.lower())] = c_id

                    cols.append(
                        ColumnModel(
                            id=c_id,
                            name=col_def.name,
                            type=col_def.data_type,
                            isPrimaryKey=col_def.is_pk,
                            isNullable=col_def.is_nullable,
                            defaultValue="",
                        )
                    )
                tables.append(TableModel(id=t_id, name=tbl_def.name, columns=cols))

            # Create relationships
            for tbl_name, tbl_def in parsed_tables.items():
                source_table_id = table_name_to_id.get(tbl_name.lower())
                if not source_table_id:
                    continue
                for local_col, ref_table, ref_col in tbl_def.fk_constraints:
                    target_table_id = table_name_to_id.get(ref_table.lower())
                    if not target_table_id:
                        continue
                    source_col_id = col_name_to_id.get(
                        (tbl_name.lower(), local_col.lower())
                    )
                    target_col_id = col_name_to_id.get(
                        (ref_table.lower(), ref_col.lower())
                    )
                    if not source_col_id or not target_col_id:
                        continue

                    rel_id = str(uuid.uuid4())[:8]
                    relationships.append(
                        RelationshipModel(
                            id=rel_id,
                            name=f"fk_{tbl_def.name}_{local_col}",
                            sourceTableId=source_table_id,
                            sourceColumnId=source_col_id,
                            targetTableId=target_table_id,
                            targetColumnId=target_col_id,
                            type="one-to-many",
                            isRequired=False,
                            cascadeDelete=False,
                            cascadeUpdate=False,
                        )
                    )

        elif file_type == "json":
            data = json.loads(text_content)
            if isinstance(data, dict) and "tables" in data:
                # Direct SchemaModel layout
                schema_model = SchemaModel.model_validate(data)
                tables = schema_model.tables
                relationships = schema_model.relationships
            elif isinstance(data, list):
                # List of objects: create a table for each
                t_id = str(uuid.uuid4())[:8]
                cols = []
                col_names = set()
                for obj in data:
                    if isinstance(obj, dict):
                        for k in obj:
                            if k not in col_names:
                                col_names.add(k)
                                cols.append(
                                    ColumnModel(
                                        id="c_" + str(uuid.uuid4())[:6],
                                        name=k,
                                        type="VARCHAR",
                                        isPrimaryKey=len(col_names) == 1,
                                        isNullable=True,
                                        defaultValue="",
                                    )
                                )
                tables.append(TableModel(id=t_id, name="imported_json", columns=cols))
            elif isinstance(data, dict):
                # Single object: create table with its key-values
                t_id = str(uuid.uuid4())[:8]
                cols = []
                for k in data:
                    cols.append(
                        ColumnModel(
                            id="c_" + str(uuid.uuid4())[:6],
                            name=k,
                            type="VARCHAR",
                            isPrimaryKey=len(cols) == 0,
                            isNullable=True,
                            defaultValue="",
                        )
                    )
                tables.append(TableModel(id=t_id, name="imported_json", columns=cols))

        elif file_type == "csv":
            f_in = io.StringIO(text_content)
            reader = csv.reader(f_in)
            headers = next(reader, [])
            t_id = str(uuid.uuid4())[:8]
            cols = []
            for h in headers:
                if h.strip():
                    cols.append(
                        ColumnModel(
                            id="c_" + str(uuid.uuid4())[:6],
                            name=h.strip(),
                            type="VARCHAR",
                            isPrimaryKey=len(cols) == 0,
                            isNullable=True,
                            defaultValue="",
                        )
                    )
            tables.append(TableModel(id=t_id, name="imported_csv", columns=cols))

        elif file_type == "xlsx":
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                for cell in ws[1]:
                    if cell.value is not None:
                        headers.append(str(cell.value).strip())
                if headers:
                    t_id = str(uuid.uuid4())[:8]
                    cols = []
                    for h in headers:
                        cols.append(
                            ColumnModel(
                                id="c_" + str(uuid.uuid4())[:6],
                                name=h,
                                type="VARCHAR",
                                isPrimaryKey=len(cols) == 0,
                                isNullable=True,
                                defaultValue="",
                            )
                        )
                    tables.append(
                        TableModel(id=t_id, name=sheet_name.lower(), columns=cols)
                    )

        elif file_type == "txt":
            # Check if it looks like DDL
            if "create table" in text_content.lower():
                from app.validation.ddl_validator import DDLValidator

                validator = DDLValidator()
                validator.validate(text_content)
                parsed_tables = validator.last_parsed_tables
                table_name_to_id = {}
                col_name_to_id = {}
                for tbl_name, tbl_def in parsed_tables.items():
                    t_id = str(uuid.uuid4())[:8]
                    table_name_to_id[tbl_name.lower()] = t_id
                    cols = []
                    for col_name, col_def in tbl_def.columns.items():
                        c_id = "c_" + str(uuid.uuid4())[:6]
                        col_name_to_id[(tbl_name.lower(), col_name.lower())] = c_id
                        cols.append(
                            ColumnModel(
                                id=c_id,
                                name=col_def.name,
                                type=col_def.data_type,
                                isPrimaryKey=col_def.is_pk,
                                isNullable=col_def.is_nullable,
                                defaultValue="",
                            )
                        )
                    tables.append(TableModel(id=t_id, name=tbl_def.name, columns=cols))
                # Relations
                for tbl_name, tbl_def in parsed_tables.items():
                    source_table_id = table_name_to_id.get(tbl_name.lower())
                    if not source_table_id:
                        continue
                    for local_col, ref_table, ref_col in tbl_def.fk_constraints:
                        target_table_id = table_name_to_id.get(ref_table.lower())
                        if not target_table_id:
                            continue
                        source_col_id = col_name_to_id.get(
                            (tbl_name.lower(), local_col.lower())
                        )
                        target_col_id = col_name_to_id.get(
                            (ref_table.lower(), ref_col.lower())
                        )
                        if not source_col_id or not target_col_id:
                            continue
                        rel_id = str(uuid.uuid4())[:8]
                        relationships.append(
                            RelationshipModel(
                                id=rel_id,
                                name=f"fk_{tbl_def.name}_{local_col}",
                                sourceTableId=source_table_id,
                                sourceColumnId=source_col_id,
                                targetTableId=target_table_id,
                                targetColumnId=target_col_id,
                                type="one-to-many",
                                isRequired=False,
                                cascadeDelete=False,
                                cascadeUpdate=False,
                            )
                        )
            else:
                # Fallback to simple comma or tab list of columns
                t_id = str(uuid.uuid4())[:8]
                cols = []
                for line in text_content.splitlines():
                    val = line.strip()
                    if val:
                        cols.append(
                            ColumnModel(
                                id="c_" + str(uuid.uuid4())[:6],
                                name=val,
                                type="VARCHAR",
                                isPrimaryKey=len(cols) == 0,
                                isNullable=True,
                                defaultValue="",
                            )
                        )
                tables.append(TableModel(id=t_id, name="imported_text", columns=cols))

        else:
            raise ValueError(f"Unsupported import file type: {file_type}")

    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse schema: {exc}",
        ) from exc

    if not tables:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No tables could be extracted from the provided input.",
        )

    # Save the newly imported schema
    schema = SchemaModel(tables=tables, relationships=relationships)

    # Save active schema to database
    tables_dict = [t.model_dump(by_alias=True) for t in schema.tables]
    relationships_dict = [r.model_dump(by_alias=True) for r in schema.relationships]

    current_schema = await db.get_active_schema(project_id=project_id)
    next_version = (current_schema["version"] + 1) if current_schema else 1

    await db.save_schema(
        project_id=project_id,
        version=next_version,
        tables=tables_dict,
        relationships=relationships_dict,
    )

    # Invalidate column guide — schema changed after import
    from app.seeder.column_guide import invalidate_column_guide

    with contextlib.suppress(Exception):
        await invalidate_column_guide(schema, runtime, db)

    return schema


async def get_schema_stats(
    _db: PersistenceProvider = Depends(get_persistence_provider),
) -> dict[str, Any]:
    """Load stats/metrics for projects, schemas, generated rows, exports, jobs, and validation."""
    import contextlib

    from sqlalchemy import text

    from app.core.settings.config import settings
    from app.llm.config_resolver import resolve_llm_config
    from app.platform.providers.sqlite_db import sqlite_db_manager

    projects_count = 0
    schemas_count = 0
    total_generated_rows = 0
    jobs_count = 0
    exports_count = 0
    validation_runs = 0
    validation_passed = 0
    validation_failed = 0

    with sqlite_db_manager.session() as s:
        # Projects Count
        with contextlib.suppress(Exception):
            projects_count = (
                s.execute(text("SELECT count(*) FROM projects")).scalar() or 0
            )

        # Schemas Count
        with contextlib.suppress(Exception):
            schemas_count = (
                s.execute(text("SELECT count(*) FROM schemas")).scalar() or 0
            )

        # Jobs Count
        with contextlib.suppress(Exception):
            jobs_count = s.execute(text("SELECT count(*) FROM jobs")).scalar() or 0

        # Exports Count
        with contextlib.suppress(Exception):
            exports_count = (
                s.execute(text("SELECT count(*) FROM export_history")).scalar() or 0
            )

        # Validation stats
        with contextlib.suppress(Exception):
            validation_runs = (
                s.execute(text("SELECT count(*) FROM validation_history")).scalar() or 0
            )
            validation_passed = (
                s.execute(
                    text(
                        "SELECT count(*) FROM validation_history WHERE result_status = 'passed'"
                    )
                ).scalar()
                or 0
            )
            validation_failed = (
                s.execute(
                    text(
                        "SELECT count(*) FROM validation_history WHERE result_status = 'failed'"
                    )
                ).scalar()
                or 0
            )

        # Total generated rows — read from dataset_metadata (authoritative persistent store)
        with contextlib.suppress(Exception):
            total_generated_rows = (
                s.execute(
                    text("SELECT COALESCE(SUM(total_rows), 0) FROM dataset_metadata")
                ).scalar()
                or 0
            )

    # --- LLM Token Analytics from llm_telemetry ---
    total_tokens = 0
    input_tokens = 0
    output_tokens = 0
    estimated_cost_usd = 0.0
    total_llm_requests = 0
    active_provider = "Unknown"
    active_model = settings.GEMINI_MODEL or "gemini-2.5-flash"

    try:
        with sqlite_db_manager.session() as s:
            row = s.execute(
                text(
                    "SELECT "
                    "  COALESCE(SUM(total_tokens), 0), "
                    "  COALESCE(SUM(prompt_tokens), 0), "
                    "  COALESCE(SUM(completion_tokens), 0), "
                    "  COALESCE(SUM(estimated_cost), 0.0), "
                    "  COUNT(*) "
                    "FROM llm_telemetry WHERE status = 'success'"
                )
            ).one()
            total_tokens = int(row[0] or 0)
            input_tokens = int(row[1] or 0)
            output_tokens = int(row[2] or 0)
            estimated_cost_usd = float(row[3] or 0.0)
            total_llm_requests = int(row[4] or 0)
    except Exception:  # noqa: S110
        pass

    with contextlib.suppress(Exception):
        llm_cfg = resolve_llm_config()
        active_provider = (llm_cfg.get("provider") or "google").capitalize()
        active_model = llm_cfg.get("model") or active_model

    tokens_per_job = (
        round(total_tokens / total_llm_requests) if total_llm_requests > 0 else 0
    )

    token_usage = {
        "total_tokens": total_tokens,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "tokens_per_job": tokens_per_job,
        "active_model": active_model,
        "active_provider": active_provider,
        "estimated_cost_usd": round(estimated_cost_usd, 6),
    }

    return {
        "projects_count": projects_count,
        "schemas_count": schemas_count,
        "total_generated_rows": total_generated_rows,
        "jobs_count": jobs_count,
        "exports_count": exports_count,
        "validation_statistics": {
            "total_runs": validation_runs,
            "passed": validation_passed,
            "failed": validation_failed,
        },
        "token_usage": token_usage,
    }
